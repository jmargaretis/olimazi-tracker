#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Schedule E reconciler — ties the .xlsx tracker to its source docs and dashboard state.

Runs locally with NO model tokens. Recomputes every Schedule E line from the raw
Income/Expenses rows (the workbook stores formulas but no cached values, so cell
values cannot be trusted), then runs a battery of bookkeeper-style checks and
writes a plain-English reconciliation report plus a machine-readable JSON the
dashboard can consume.

Usage:
    python _System/reconcile.py            # writes report + json, prints summary
    python _System/reconcile.py --quiet    # just exit code (0 = all PASS)

Exit code is the number of FAIL checks (0 = fully reconciled).
"""
from __future__ import annotations
import glob
import json
import os
import re
import sys
from datetime import date
from html import escape

import openpyxl

# ---- config -------------------------------------------------------------
# Multi-property: defaults to this project (12 Sample Street), but
# `python reconcile.py <other-project-root>` reconciles any property folder
# that holds a *Tracker*.xlsx with the same tab layout.
#
# ROOT is PINNED to the live OneDrive workspace (Claude OS) so the reconciler
# always reads and writes the canonical Sch. E folder — never the frozen Google
# Drive backup — even if a stray copy of this script is run from elsewhere.
# Falls back to the script-relative location only if that path doesn't exist
# (e.g. the workspace was relocated to another machine).
PINNED_ROOT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "fixtures", "sample-property")
DEFAULT_ROOT = PINNED_ROOT if os.path.isdir(PINNED_ROOT) else \
    os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def configure(root):
    global ROOT, TRACKER, DASH_STATE, REPORT, JSON_OUT, DASHBOARD
    ROOT = os.path.abspath(root)
    # The tracker moved to the project root (sample-tracker.xlsx) on 2026-07-16;
    # search both the project root and ROOT so either location works.
    cands = [f for f in glob.glob(os.path.join(os.path.dirname(ROOT), "*.xlsx")) +
             glob.glob(os.path.join(ROOT, "*.xlsx"))
             if "backup" not in os.path.basename(f).lower()
             and "test" not in os.path.basename(f).lower()]
    pref = [f for f in cands if "tracker" in os.path.basename(f).lower()]
    TRACKER = (pref or cands or [os.path.join(os.path.dirname(ROOT), "sample-tracker.xlsx")])[0]
    DASH_STATE = os.path.join(ROOT, "dashboard_state.json")
    # Fixed filename (overwrite in place) — no date in the name, so re-running never
    # leaves a trail of dated report copies. The run date is recorded inside the file.
    REPORT = os.path.join(ROOT, "SchE_Reconciliation.md")
    JSON_OUT = os.path.join(ROOT, "reconciliation.json")
    DASHBOARD = os.path.join(ROOT, "SchE_Dashboard.html")


configure(DEFAULT_ROOT)

CENTS = 0.005  # tolerance for float compares

# Client-side logic for the dashboard's "Anticipated" card. Computes the deadline
# countdown LIVE on each page open (so the 30-day red trigger reflects today, not the
# last reconcile run) and responds to the extension toggle. Kept as a plain string so
# its braces don't collide with the dashboard f-string.
# Write-back UI: buttons stay hidden when the dashboard is opened as a plain file;
# they activate only when served by _System/dashboard_server.py (localhost).
RESOLVE_JS = r"""
(function(){
  if (!location.protocol.startsWith('http')) return;
  document.querySelectorAll('.resolvebtn').forEach(function(btn){
    btn.hidden = false;
    btn.addEventListener('click', function(){
      var row = btn.getAttribute('data-row');
      var title = btn.closest('.chip').querySelector('.chip-t').childNodes[0].textContent.trim();
      var note = prompt('Resolve "' + title + '" — add a short note (or leave blank):');
      if (note === null) return;
      btn.disabled = true; btn.textContent = 'Saving…';
      fetch('/api/resolve', {
        method: 'POST', headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({row: Number(row), note: note})
      }).then(function(r){ return r.json(); }).then(function(j){
        if (j.ok) { location.reload(); }
        else { alert('Not saved: ' + j.error); btn.disabled = false; btn.textContent = 'Resolve'; }
      }).catch(function(e){ alert('Server error: ' + e); btn.disabled = false; btn.textContent = 'Resolve'; });
    });
  });
})();
"""

ANT_JS = r"""
(function(){
  if(typeof TAX==='undefined'||!TAX) return;
  var chk=document.getElementById('extchk');
  var KEY='sche_ext_filed_'+(TAX.tax_year||'');
  if(chk) chk.checked = localStorage.getItem(KEY)==='1';
  function fmt(d){return d.toLocaleDateString(undefined,{year:'numeric',month:'short',day:'numeric'});}
  function render(){
    var ext = chk && chk.checked;
    var dl = new Date((ext?TAX.extension_deadline:TAX.deadline)+'T00:00:00');
    var today=new Date(); today.setHours(0,0,0,0);
    var days=Math.round((dl-today)/86400000);
    var bar=document.getElementById('deadbar');
    var which = ext?'Extended deadline':'Filing deadline';
    var urg = days<=30?'red':(days<=60?'amber':'ok');
    if(bar){ bar.className='deadbar '+urg;
      var extTxt = ext?' · extension ACTIVE':(' · extension available: '+fmt(new Date(TAX.extension_deadline+'T00:00:00')));
      bar.innerHTML = which+': <b>'+fmt(dl)+'</b> · '+(days>=0?days+' days left':(Math.abs(days)+' days OVERDUE'))+extTxt; }
    var rows=document.querySelectorAll('tr[data-status]');
    for(var i=0;i<rows.length;i++){
      var tr=rows[i], st=tr.getAttribute('data-status'), cell=tr.querySelector('.stat');
      if(!cell) continue;
      if(st==='booked'){ cell.innerHTML='<span class=pill style="background:#566123">booked</span>'; tr.style.background=''; tr.style.opacity='1'; }
      else if(st==='na'){ cell.innerHTML='<span class=pill style="background:#9a917c">n/a</span>'; tr.style.background=''; tr.style.opacity='.6'; }
      else { var red=days<=30; cell.innerHTML='<span class=pill style="background:'+(red?'#9B1C2E':'#9c6f19')+'">'+(red?'MISSING':'expected')+'</span>'; tr.style.background=red?'#f6e0dd':'#f3e6c7'; tr.style.opacity='1'; }
    }
  }
  if(chk) chk.addEventListener('change',function(){ localStorage.setItem(KEY, chk.checked?'1':'0'); render(); });
  render();
})();
"""


def money(x):
    return f"${x:,.0f}"


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def line_no(label):
    """Extract the Schedule E line number from a label like 'L14 — Repairs'."""
    if not isinstance(label, str):
        return None
    m = re.search(r"L\s*(\d+)", label)
    return int(m.group(1)) if m else None


# ---- load ---------------------------------------------------------------
def load():
    wb = openpyxl.load_workbook(TRACKER, data_only=False)
    return wb


def read_profile(wb):
    """Read optional setup metadata without modifying the tracker or its sources."""
    fallback_name = str(wb["Income"]["B4"].value or "Unnamed tracker").strip()
    profile = {
        "name": fallback_name,
        "type": "rental",
        "applicable_lines": list(range(3, 20)),
        "engine_support": "schedule_e",
    }
    path = os.path.join(ROOT, "profile.json")
    if os.path.exists(path):
        try:
            loaded = json.load(open(path, encoding="utf-8"))
            if isinstance(loaded, dict):
                profile.update(loaded)
        except (OSError, ValueError):
            pass
    profile["name"] = str(profile.get("name") or fallback_name).replace("\n", " ").strip()
    profile["type"] = "business" if profile.get("type") == "business" else "rental"
    profile["applicable_lines"] = [
        line for line in profile.get("applicable_lines", []) if isinstance(line, int)
    ]
    return profile


def read_income(ws):
    """Return (ytd_net, full_year_net, actual_net, rows)."""
    rows = []
    for r in range(6, 18):  # Jan..Dec
        month = ws.cell(r, 1).value
        if not month:
            continue
        rent = num(ws.cell(r, 2).value)
        credit = num(ws.cell(r, 3).value)
        status = ws.cell(r, 5).value  # column E = Status (D is the Net Rent formula)
        rows.append({"month": month, "rent": rent, "credit": credit,
                     "net": rent - credit, "status": str(status or "")})
    actual = [x for x in rows if "Actual" in x["status"]]
    ytd_net = sum(x["net"] for x in rows[:4])          # Jan-Apr (Summary links D19)
    full_net = sum(x["net"] for x in rows)
    actual_net = sum(x["net"] for x in actual)
    return ytd_net, full_net, actual_net, rows


def read_expenses(ws):
    rows = []
    for r in range(5, ws.max_row + 1):
        date_v = ws.cell(r, 1).value
        vendor = ws.cell(r, 2).value
        if (vendor in (None, "TOTAL")) and date_v in (None, "TOTAL"):
            continue
        rows.append({
            "row": r,
            "date": str(date_v) if date_v else "",
            "vendor": vendor or "",
            "desc": ws.cell(r, 3).value or "",
            "line_label": ws.cell(r, 4).value or "",
            "line": line_no(ws.cell(r, 4).value),
            "contracted": num(ws.cell(r, 5).value),
            "paid": num(ws.cell(r, 6).value),
            "balance": num(ws.cell(r, 7).value),
            "status": ws.cell(r, 8).value or "",
            "file_ref": ws.cell(r, 9).value or "",
            "notes": ws.cell(r, 10).value or "",
        })
    return rows


# ---- checks -------------------------------------------------------------
class Report:
    def __init__(self):
        self.checks = []

    def add(self, name, ok, detail, severity="fail"):
        self.checks.append({"name": name, "status": "PASS" if ok else severity.upper(),
                            "detail": detail})

    @property
    def fails(self):
        return [c for c in self.checks if c["status"] == "FAIL"]

    @property
    def warns(self):
        return [c for c in self.checks if c["status"] == "WARN"]


def main():
    quiet = "--quiet" in sys.argv
    roots = [a for a in sys.argv[1:] if not a.startswith("--")]
    if roots:
        configure(roots[0])

    # Guard: delete any stray dashboard_state.json sitting in the workspace ROOT.
    # The only copy we keep lives inside _System\ (a different path), so this removes
    # accidental duplicates at the folder root without touching the kept one.
    stray_state = os.path.join(ROOT, "dashboard_state.json")
    if os.path.isfile(stray_state):
        try:
            os.remove(stray_state)
            if not quiet:
                print(f"Removed stray {stray_state}")
        except OSError as e:
            print(f"Could not remove stray {stray_state}: {e}")

    if not os.path.exists(TRACKER):
        print(f"Tracker not found: {TRACKER}")
        return 99
    wb = load()
    profile = read_profile(wb)
    inc = wb["Income"]
    exp = wb["Expenses"]

    ytd_net, full_net, actual_net, inc_rows = read_income(inc)
    erows = read_expenses(exp)
    rep = Report()

    # --- recompute Schedule E lines from raw rows ---
    by_line = {}
    for e in erows:
        if e["line"] is not None:
            by_line.setdefault(e["line"], 0.0)
            by_line[e["line"]] += e["paid"]
    line9 = by_line.get(9, 0.0)
    line10 = by_line.get(10, 0.0)
    line14 = by_line.get(14, 0.0)
    total_exp = sum(by_line.values())
    line3 = ytd_net
    net_income = line3 + 0.0 - total_exp  # royalties (L4) = 0
    anticipated = read_anticipated(by_line)

    # 1) Income internal tie: YTD net == sum of Actual months
    rep.add("Income: YTD (Jan-Apr) == sum of 'Actual' months",
            abs(ytd_net - actual_net) < CENTS,
            f"YTD net {money(ytd_net)} vs Actual-flagged months {money(actual_net)}. "
            f"Summary Line 3 links to Income!D19 (Jan-Apr).")

    # 2) Expenses TOTAL formula target == recomputed
    paid_total = sum(e["paid"] for e in erows)
    rep.add("Expenses: category lines sum to Paid-YTD total",
            abs(total_exp - paid_total) < CENTS,
            f"Sum of L-line categories {money(total_exp)} vs sum of all Paid-YTD {money(paid_total)}.")

    # 3) SUMIF fragility: any row whose label won't exactly match the canonical strings
    canon = {9: "L9 — Insurance", 10: "L10 — Professional fees", 14: "L14 — Repairs"}
    fragile = []
    for e in erows:
        if e["line"] in canon and e["line_label"].strip() != canon[e["line"]]:
            fragile.append(f"row {e['row']}: '{e['line_label']}' != '{canon[e['line']]}'")
    rep.add("Excel SUMIF strings match canonical labels (em-dash exact)",
            not fragile,
            "All expense labels match the exact SUMIF text." if not fragile
            else "Excel SUMIF would DROP these rows from the rollup: " + "; ".join(fragile),
            severity="warn")

    # 4) File reference integrity
    missing = []
    for e in erows:
        ref = str(e["file_ref"]).strip()
        if not ref:
            continue
        p = os.path.join(ROOT, ref.replace("/", os.sep))
        if not os.path.exists(p):
            missing.append(f"row {e['row']} ({e['vendor']}): {ref}")
    rep.add("File references resolve to real files on disk",
            not missing,
            "Every File Reference points to an existing file." if not missing
            else "Dangling File References (the document isn't where the tracker says): "
                 + " | ".join(missing))

    # 5) Insurance-reimbursement reconciliation (the headline)
    reimb = detect_reimbursement(wb)
    if reimb["amount"] > 0 and not reimb["booked"]:
        net_correct = net_income + reimb["amount"]
        claim_repairs = sum(e["paid"] for e in erows
                            if e["line"] == 14 and ("Claim" in str(e["desc"]) or "Claim" in str(e["notes"])
                                                    or "SAMPLE-001" in str(e["notes"]) or "REF-0000" in str(e["notes"])))
        net_oop = claim_repairs - reimb["amount"]
        rep.add("Insurance reimbursement is booked",
                False,
                f"Reimbursement of {money(reimb['amount'])} (Claim SAMPLE-001) is recorded as RECEIVED "
                f"in the Review tab but is NOT booked in Expenses/Summary. "
                f"Claim repairs booked at gross {money(claim_repairs)}; net out-of-pocket should be "
                f"{money(net_oop)}. Net income as shown {money(net_income)} understates true net by "
                f"{money(reimb['amount'])} -> corrected Net income {money(net_correct)}. "
                f"Fix: net the proceeds against Line 14 (or book as income) before filing.")
    else:
        rep.add("Insurance reimbursement is booked",
                reimb["amount"] == 0 or reimb["booked"],
                "No outstanding received-but-unbooked reimbursement detected.")

    # 6) Open balances on the Expenses tab
    open_bal = [e for e in erows if abs(e["balance"]) > CENTS]
    rep.add("No unexplained open balances on Expenses",
            not open_bal,
            "All contracted amounts fully paid (Balance = 0)." if not open_bal
            else "Open balances: " + "; ".join(f"row {e['row']} {e['vendor']} {money(e['balance'])}" for e in open_bal),
            severity="warn")

    # 7) Items still TBD / unallocated (filing-readiness, informational)
    tbd = [e for e in erows if "TBD" in str(e["status"]).upper() or "TBD" in str(e["contracted"]).upper()]
    rep.add("No unallocated/TBD expense rows",
            not tbd,
            "No TBD expense rows." if not tbd
            else "TBD/unallocated: " + "; ".join(f"row {e['row']} {e['vendor']}" for e in tbd),
            severity="warn")

    # 8) Accountant-review queue — any row explicitly flagged for a preparer
    acct_flags = collect_accountant_flags(wb, erows)
    rep.add("Accountant-review queue is clear",
            not acct_flags,
            "No items flagged for accountant review." if not acct_flags
            else f"{len(acct_flags)} item(s) flagged for a preparer to confirm: "
                 + " | ".join(a["item"] for a in acct_flags),
            severity="warn")

    # 9) Anticipated prior-year items not yet booked this year (filing-readiness)
    if anticipated:
        miss = [it for it in anticipated["items"] if it["status"] == "missing"]
        rep.add("Anticipated prior-year items are booked this year",
                not miss,
                "All anticipated items have a current-year amount." if not miss
                else "Expected from prior year, not yet booked: " + "; ".join(
                    f"{m['label']} (PY {m['py_year']}: "
                    f"{money(m['py_amount']) if isinstance(m['py_amount'], (int, float)) else 'n/a'})"
                    for m in miss),
                severity="warn")

    open_items = read_review_open(wb)
    summary = {
        "generated": str(date.today()),
        "profile": profile,
        "property_name": profile["name"],
        "tracking_type": profile["type"],
        "applicable_lines": profile["applicable_lines"],
        "tracker_name": os.path.basename(TRACKER),
        "open_items": open_items,
        "anticipated": anticipated,
        "accountant_review": acct_flags,
        "line3_rents_ytd": round(line3, 2),
        "line9_insurance": round(line9, 2),
        "line10_professional": round(line10, 2),
        "line14_repairs": round(line14, 2),
        "total_expenses": round(total_exp, 2),
        "net_income_as_booked": round(net_income, 2),
        "reimbursement": reimb,
        "net_income_corrected": round(net_income + (reimb["amount"] if not reimb["booked"] else 0), 2),
        "fails": len(rep.fails),
        "warns": len(rep.warns),
        "checks": rep.checks,
    }
    write_report(summary, rep)
    write_dashboard(summary, rep)
    with open(JSON_OUT, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)

    if not quiet:
        print(f"RECONCILIATION  {summary['generated']}")
        print(f"  Net income as booked : {money(summary['net_income_as_booked'])}")
        print(f"  Net income corrected : {money(summary['net_income_corrected'])}")
        print(f"  Checks: {len(rep.checks)}  FAIL={len(rep.fails)}  WARN={len(rep.warns)}")
        for c in rep.checks:
            mark = {"PASS": "[ok]", "FAIL": "[FAIL]", "WARN": "[warn]"}[c["status"]]
            print(f"  {mark} {c['name']}")
        print(f"  Report: {REPORT}")
    return len(rep.fails)


def detect_reimbursement(wb):
    """Scan Review tab for a received-but-unbooked insurance reimbursement."""
    amount, booked, evidence = 0.0, False, ""
    if "Review" in wb.sheetnames:
        ws = wb["Review"]
        for r in range(1, ws.max_row + 1):
            txt = " ".join(str(ws.cell(r, c).value or "") for c in range(1, 6))
            if "reimbursement" in txt.lower() and "received" in txt.lower():
                # grab the amount tied to "received", not the first $ in the cell
                m = re.search(r"received[^$]*\$([\d,]+\.\d{2})", txt, re.IGNORECASE)
                if not m:  # fallback: largest dollar amount in the cell
                    amts = [float(a.replace(",", "")) for a in re.findall(r"\$([\d,]+\.\d{2})", txt)]
                    if amts:
                        amount = max(amts)
                        evidence = txt.strip()[:200]
                    continue
                amount = float(m.group(1).replace(",", ""))
                evidence = txt.strip()[:200]
    # "booked" would require an income line or a contra row; none exists in this layout
    exp = wb["Expenses"]
    for r in range(5, exp.max_row + 1):
        v = " ".join(str(exp.cell(r, c).value or "") for c in (2, 3, 4))
        if "reimburs" in v.lower() or "proceeds" in v.lower():
            booked = True
    return {"amount": round(amount, 2), "booked": booked, "evidence": evidence}


ACCT_MARKER = "ACCOUNTANT REVIEW"


def collect_accountant_flags(wb, erows):
    """Standing queue of items a preparer should confirm.

    Picks up (a) any Expenses Notes cell containing the 'ACCOUNTANT REVIEW' marker,
    so the system can flag similar items going forward just by adding that phrase to a
    row's note, and (b) capitalize-vs-expense / claim items that always warrant a look.
    """
    flags = []
    for e in erows:
        note = str(e["notes"])
        if ACCT_MARKER.lower() in note.lower():
            flags.append({"where": f"Expenses row {e['row']}",
                          "item": f"{e['vendor']} — {money(e['paid'])}",
                          "note": note.split(ACCT_MARKER, 1)[-1].lstrip(": ").strip()[:240]})
    if "Review" in wb.sheetnames:
        ws = wb["Review"]
        for r in range(5, ws.max_row + 1):
            txt = " ".join(str(ws.cell(r, c).value or "") for c in range(1, 6))
            if ACCT_MARKER.lower() in txt.lower():
                flags.append({"where": f"Review row {r}",
                              "item": str(ws.cell(r, 2).value or ""),
                              "note": str(ws.cell(r, 4).value or "")[:240]})
    return flags


def read_review_open(wb):
    """Collect open (non-resolved) items from the Review tab for dashboard chips."""
    items = []
    if "Review" not in wb.sheetnames:
        return items
    ws = wb["Review"]
    for r in range(5, ws.max_row + 1):
        typ = ws.cell(r, 1).value
        item = ws.cell(r, 2).value
        status = str(ws.cell(r, 5).value or "")
        if not item:
            continue
        resolved = status.lower().startswith("resolved")
        items.append({
            "type": str(typ or ""), "item": str(item),
            "action": str(ws.cell(r, 4).value or ""),
            "status": status, "resolved": resolved, "row": r,
        })
    return items


def read_anticipated(by_line):
    """Prior-year line items expected to recur this year, from ROOT/anticipated.json
    (per property). Each item is compared to the current-year booked amount so the
    dashboard can show it booked (green) or still-expected (amber, with the PY amount
    as a placeholder). Returns None if no config exists."""
    cfg_path = os.path.join(ROOT, "anticipated.json")
    if not os.path.exists(cfg_path):
        return None
    try:
        cfg = json.load(open(cfg_path, encoding="utf-8"))
    except Exception:
        return None
    items = []
    for it in cfg.get("items", []):
        ln = it.get("line")
        cy = round(by_line.get(ln, 0.0), 2) if isinstance(ln, int) else 0.0
        na = bool(it.get("not_applicable"))
        items.append({
            "line": ln, "label": it.get("label", ""),
            "py_amount": it.get("py_amount"), "py_year": it.get("py_year"),
            "note": it.get("note", ""), "reason": it.get("reason", ""),
            "cy_amount": cy,
            "status": "na" if na else ("booked" if cy > CENTS else "missing"),
        })
    return {"tax_year": cfg.get("tax_year"), "deadline": cfg.get("deadline"),
            "extension_deadline": cfg.get("extension_deadline"), "items": items}


def write_dashboard(s, rep):
    """Emit a self-contained HTML dashboard generated FROM the reconciliation, so it
    can never drift from the spreadsheet."""
    fails = [c for c in rep.checks if c["status"] == "FAIL"]
    warns = [c for c in rep.checks if c["status"] == "WARN"]
    verdict = "RECONCILED" if not fails else f"{len(fails)} ALERT(S)"
    vcolor = "#566123" if not fails else "#9B1C2E"
    open_unres = [i for i in s.get("open_items", []) if not i["resolved"]]
    schedule_label = (
        "Schedule E" if s["tracking_type"] == "rental" else "Schedule C (setup preview)"
    )
    property_name = escape(s["property_name"])
    tracker_name = escape(s["tracker_name"])
    applicable = set(s.get("applicable_lines", []))

    def chips():
        out = []
        if s["net_income_corrected"] != s["net_income_as_booked"]:
            out.append(("#9B1C2E", "Deduction at risk",
                        f"Reimbursement {money(s['reimbursement']['amount'])} unbooked → "
                        f"net should be {money(s['net_income_corrected'])}", None))
        for i in open_unres:
            color = "#9c6f19" if i["type"].lower() in ("missing data", "insurance allocation") else "#8a6d3b"
            out.append((color, i["item"], i["action"][:90], i.get("row")))
        if not out:
            out.append(("#566123", "No open items", "Everything reconciles", None))
        return out

    def dashboard_row(label, value):
        line = int(label.split()[0])
        inactive = line not in applicable and line != 20
        tag = " <span class=tag>not selected</span>" if inactive else ""
        css_class = "inactive" if inactive else ""
        return (
            f"<tr class='{css_class}'><td>{label}{tag}</td>"
            f"<td class='num'>{money(value)}</td></tr>"
        )

    rows = "\n".join(
        dashboard_row(lbl, val)
        for lbl, val in [
            ("3 — Rents received (YTD)", s["line3_rents_ytd"]),
            ("9 — Insurance", s["line9_insurance"]),
            ("10 — Professional fees", s["line10_professional"]),
            ("14 — Repairs", s["line14_repairs"]),
            ("20 — Total expenses", s["total_expenses"]),
        ])
    corrected_row = ""
    if s["net_income_corrected"] != s["net_income_as_booked"]:
        corrected_row = (f"<tr class='corrected'><td>21 — Net income "
                         f"<span class='tag'>corrected for reimbursement</span></td>"
                         f"<td class='num'>{money(s['net_income_corrected'])}</td></tr>")
    chip_html = "\n".join(
        f"<div class='chip'{f' data-row={r}' if r else ''} style='border-left:5px solid {c}'>"
        f"<div class='chip-t'>{t}"
        + (f" <button class='resolvebtn' data-row='{r}' hidden>Resolve</button>" if r else "")
        + f"</div><div class='chip-d'>{d}</div></div>"
        for c, t, d, r in chips())
    acct = s.get("accountant_review", [])
    if acct:
        acct_html = "\n".join(
            f"<div class='chip' style='border-left:5px solid #8a6d3b'>"
            f"<div class='chip-t'>{a['item']} <span style='color:#9a917c;font-weight:400'>· {a['where']}</span></div>"
            f"<div class='chip-d'>{a['note']}</div></div>" for a in acct)
    else:
        acct_html = ("<div class='chip-d'>None. Add <code>ACCOUNTANT REVIEW:</code> to any tracker "
                     "Note to surface it here.</div>")
    check_html = "\n".join(
        f"<li class='{c['status'].lower()}'>"
        f"<span class='ic'>{'✅' if c['status']=='PASS' else '❌' if c['status']=='FAIL' else '⚠️'}</span>"
        f"<b>{c['name']}</b><br><span class='cd'>{c['detail']}</span></li>"
        for c in rep.checks)

    ant = s.get("anticipated")
    anticipated_css = anticipated_html = anticipated_script = ""
    if ant and ant.get("items"):
        item_rows = ""
        for it in ant["items"]:
            py = it["py_amount"]
            py_txt = money(py) if isinstance(py, (int, float)) else "n/a"
            cy_txt = (money(it["cy_amount"]) if it["status"] == "booked"
                      else "n/a" if it["status"] == "na" else "—")
            lbl = (f"{it['line']} — " if it["line"] else "") + it["label"]
            note_txt = it.get("reason") or it["note"]
            item_rows += (
                f"<tr data-status='{it['status']}'>"
                f"<td>{lbl}</td><td class='num'>{cy_txt}</td>"
                f"<td class='py num'>PY {it['py_year'] or ''}: {py_txt}</td>"
                f"<td class='stat'></td><td class='note'>{note_txt}</td></tr>")
        anticipated_css = (
            ".deadbar{padding:9px 13px;border-radius:10px;font-size:13px;margin-bottom:12px;background:#efe7d6;color:#5b5444}"
            ".deadbar.ok{background:#e7ecd6;color:#404a18}.deadbar.amber{background:#f3e6c7;color:#8a5a12}"
            ".deadbar.red{background:#f6e0dd;color:#7d1624;font-weight:700}"
            ".exttoggle{display:inline-flex;align-items:center;gap:6px;font-size:13px;margin:0 0 12px;cursor:pointer;color:#5b5444}"
            ".anttable{width:100%;border-collapse:collapse;font-size:13px}"
            ".anttable th{text-align:left;color:#7d7666;font-weight:700;border-bottom:2px solid #D2C8B0;padding:6px 4px}"
            ".anttable td{padding:7px 4px;border-bottom:1px solid #e7ddc8;vertical-align:top}"
            ".anttable .py{color:#7d7666}.anttable .note{color:#9a917c;font-size:12px}"
            ".pill{color:#fff;border-radius:999px;padding:1px 9px;font-size:11px;font-weight:700}"
            ".legend{color:#9a917c;font-size:11px;margin-top:8px}")
        anticipated_html = (
            "<div class=card style=\"margin-top:18px\" id=anticipated>"
            "<h2>Anticipated this year (from prior year)</h2>"
            "<div id=deadbar class=deadbar></div>"
            "<label class=exttoggle><input type=checkbox id=extchk> Extension filed with the IRS</label>"
            "<table class=anttable><thead><tr><th>Line</th><th class=num>CY booked</th>"
            "<th class=num>Prior year</th><th>Status</th><th>Note</th></tr></thead>"
            f"<tbody>{item_rows}</tbody></table>"
            "<div class=legend>Amber = expected, not yet provided &middot; Red = still missing within 30 days of the deadline &middot; Green = booked. Toggle the extension box to shift the deadline.</div></div>")
        tax_data = json.dumps({"deadline": ant.get("deadline"),
                               "extension_deadline": ant.get("extension_deadline"),
                               "tax_year": ant.get("tax_year")})
        anticipated_script = "<script>\nvar TAX=" + tax_data + ";\n" + ANT_JS + "\n</script>"

    html = f"""<!doctype html><html lang=en><head><meta charset=utf-8>
<meta name=viewport content="width=device-width,initial-scale=1">
<title>{schedule_label} — {property_name}</title>
<link rel="icon" href="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 64 64'><rect width='64' height='64' rx='14' fill='%23566123'/><path d='M12 34 L32 18 L52 34' fill='none' stroke='%23F4EDDD' stroke-width='5'/><path d='M16 46 H48' stroke='%23F4EDDD' stroke-width='4'/><path d='M16 54 H40' stroke='%239B1C2E' stroke-width='4'/></svg>">
<style>
:root{{
  --cream:#F4EDDD;--card:#FBF8EF;--crimson:#9B1C2E;--olive:#566123;
  --sand:#D2C8B0;--ink:#2B2A26;--muted:#7D7666;
  font-family:'Inter','Segoe UI',-apple-system,Roboto,Helvetica,Arial,sans-serif;
}}
*{{box-sizing:border-box}}
body{{margin:0 auto;background:var(--cream);color:var(--ink);padding:26px 20px;max-width:1040px;-webkit-font-smoothing:antialiased}}
nav{{display:flex;align-items:center;justify-content:space-between;gap:12px;padding:13px 18px;background:var(--card);border:1px solid var(--sand);border-radius:14px;box-shadow:0 1px 2px rgba(43,42,38,.06)}}
.brand{{display:flex;align-items:center;gap:11px}}
.brand .mark{{flex:0 0 auto;display:block}}
.wm{{font-size:21px;font-weight:800;letter-spacing:-.02em;color:var(--ink)}}
.wm .wd{{color:var(--muted);font-weight:500}}
.right{{display:flex;align-items:center;gap:12px}} .stamp{{color:var(--muted);font-size:12px}}
.verdict{{display:inline-block;padding:6px 14px;border-radius:999px;color:#fff;font-weight:700;font-size:13px;background:{vcolor}}}
.hero{{padding:28px 6px 6px}}
.hero h1{{font-size:34px;line-height:1.05;margin:0;font-weight:800;letter-spacing:-.02em}}
.hero .addr{{margin:6px 0 4px;font-size:15px;font-weight:600;color:var(--ink)}}
.hero p{{margin:0;color:var(--muted);font-size:13px;max-width:640px}}
.grid{{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-top:16px}}
@media(max-width:760px){{.grid{{grid-template-columns:1fr}}.hero h1{{font-size:27px}}}}
.card{{background:var(--card);border:1px solid var(--sand);border-radius:16px;padding:18px;box-shadow:0 1px 2px rgba(43,42,38,.06)}}
.card h2{{font-size:12px;margin:0 0 12px;text-transform:uppercase;letter-spacing:.08em;color:var(--olive);font-weight:700}}
table{{width:100%;border-collapse:collapse;font-size:14px}}
td{{padding:8px 4px;border-bottom:1px solid var(--sand)}} .num{{text-align:right;font-variant-numeric:tabular-nums}}
tr:last-child td{{border-bottom:none}}
.corrected td{{font-weight:700;color:var(--crimson)}} .tag{{font-size:10px;background:#f1e3d2;color:var(--crimson);padding:1px 6px;border-radius:6px;font-weight:600}}
.booked td{{font-weight:800}}
.inactive{{opacity:.42}}
.chip{{background:#fff;border:1px solid var(--sand);border-radius:12px;padding:10px 12px;margin-bottom:10px}}
.chip-t{{font-weight:700;font-size:14px}} .chip-d{{font-size:12px;color:var(--muted);margin-top:2px}}
ul.checks{{list-style:none;padding:0;margin:0;font-size:13px}}
ul.checks li{{padding:11px 0;border-bottom:1px solid var(--sand);display:flex;gap:8px;flex-wrap:wrap}}
ul.checks li:last-child{{border-bottom:none}}
ul.checks li .ic{{flex:0 0 auto}} ul.checks li b{{flex:1 1 auto}}
.cd{{color:var(--muted);width:100%;padding-left:26px}}
li.fail b{{color:var(--crimson)}}
code{{background:#efe7d6;padding:1px 5px;border-radius:5px;font-size:.9em}}
.resolvebtn{{float:right;background:var(--olive);color:#fff;border:none;border-radius:8px;padding:3px 10px;font-size:11px;font-weight:700;cursor:pointer}}
.resolvebtn:hover{{background:#455018}} .resolvebtn:disabled{{opacity:.6;cursor:wait}}
footer{{color:var(--muted);font-size:12px;margin-top:24px;text-align:center}}
{anticipated_css}
</style></head><body>
<nav>
  <div class=brand>
    <svg class="mark" width="42" height="42" viewBox="0 0 64 64" aria-hidden="true"><rect width="64" height="64" rx="16" fill="#566123"></rect><path d="M12 34 L32 18 L52 34" fill="none" stroke="#F4EDDD" stroke-width="5" stroke-linecap="round" stroke-linejoin="round"></path><path d="M16 46 H48" fill="none" stroke="#F4EDDD" stroke-width="4" stroke-linecap="round"></path><path d="M16 54 H40" fill="none" stroke="#9B1C2E" stroke-width="4" stroke-linecap="round"></path></svg>
    <span class=wm>olimazi<span class=wd>.online</span></span>
  </div>
  <div class=right><span class=stamp>Updated {s['generated']}</span><span class=verdict>{verdict}</span></div>
</nav>
<div class=hero>
  <h1>{schedule_label}</h1>
  <p class=addr>{property_name}</p>
  <p>Reconciled from <code>{tracker_name}</code> &middot; deterministic arithmetic &middot; always matches the spreadsheet.</p>
</div>
<div class=grid>
  <div class=card><h2>{schedule_label} bottom line</h2>
    <table>{rows}
      <tr class=booked><td>21 — Net income (as booked)</td><td class=num>{money(s['net_income_as_booked'])}</td></tr>
      {corrected_row}
      <tr class=booked><td>Grand total (L3 − L20)</td><td class=num>{money(s['line3_rents_ytd'] - s['total_expenses'])}</td></tr>
    </table></div>
  <div class=card><h2>Open items ({len(open_unres)})</h2>{chip_html}</div>
</div>
{anticipated_html}
<div class=card style="margin-top:18px"><h2>Flagged for accountant review ({len(s.get('accountant_review', []))})</h2>
  {acct_html}</div>
<div class=card style="margin-top:18px"><h2>Reconciliation checks</h2>
  <ul class=checks>{check_html}</ul></div>
<footer>Re-run: <code>python _System/reconcile.py</code> · {len(fails)} alert(s), {len(warns)} warning(s)</footer>
{anticipated_script}
<script>{RESOLVE_JS}</script>
</body></html>"""
    with open(DASHBOARD, "w", encoding="utf-8") as f:
        f.write(html)


def write_report(s, rep):
    L = []
    schedule_label = (
        "Schedule E" if s["tracking_type"] == "rental" else "Schedule C (setup preview)"
    )
    L.append(f"# {schedule_label} — Reconciliation Report")
    L.append(f"**Tracker:** {s['property_name']} · **Generated:** {s['generated']} · "
             f"_Deterministic check — recomputed from raw rows, no model judgment._\n")
    verdict = "RECONCILED" if s["fails"] == 0 else f"{s['fails']} ALERT(S) FOUND"
    L.append(f"## Verdict: **{verdict}**  ({s['warns']} warning(s))\n")
    L.append("## Bottom line")
    L.append(f"| Line | Amount |")
    L.append(f"|---|---|")
    L.append(f"| 3 — Rents received (YTD) | {money(s['line3_rents_ytd'])} |")
    L.append(f"| 9 — Insurance | {money(s['line9_insurance'])} |")
    L.append(f"| 10 — Professional fees | {money(s['line10_professional'])} |")
    L.append(f"| 14 — Repairs | {money(s['line14_repairs'])} |")
    L.append(f"| 20 — Total expenses | {money(s['total_expenses'])} |")
    L.append(f"| **21 — Net income (as booked)** | **{money(s['net_income_as_booked'])}** |")
    if s["net_income_corrected"] != s["net_income_as_booked"]:
        L.append(f"| **21 — Net income (corrected for reimbursement)** | **{money(s['net_income_corrected'])}** |")
    L.append(f"| **Grand total (L3 − L20)** | **{money(s['line3_rents_ytd'] - s['total_expenses'])}** |")
    L.append("")
    L.append("## Checks")
    for c in rep.checks:
        icon = {"PASS": "✅", "FAIL": "❌", "WARN": "⚠️"}[c["status"]]
        L.append(f"- {icon} **{c['name']}** — {c['detail']}")
    L.append("")
    acct = s.get("accountant_review", [])
    L.append("## Flagged for accountant review")
    if not acct:
        L.append("_None. Add the phrase `ACCOUNTANT REVIEW:` to any tracker Note to surface it here._")
    else:
        for a in acct:
            L.append(f"- **{a['item']}** ({a['where']}) — {a['note']}")
    L.append("")
    L.append("_Re-run any time with `python _System/reconcile.py`. Exit code = number of alerts._")
    with open(REPORT, "w", encoding="utf-8") as f:
        f.write("\n".join(L))


if __name__ == "__main__":
    sys.exit(main())
