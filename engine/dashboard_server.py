#!/usr/bin/env python3
"""Local write-back server for the Schedule E dashboard.

Serves SchE_Dashboard.html on http://127.0.0.1:8742 and accepts one write:
POST /api/resolve {row, note} -> sets Review!E<row> to "Resolved — <note> (<date>)",
then re-runs reconcile.py so the dashboard/report regenerate from the books.

Design rules (match the system's hard rules):
- Localhost only. The ONLY cell this server ever writes is the Status column (E)
  of an OPEN row on the Review tab — a click in the UI is John's explicit
  confirmation for that one write. Nothing else in the workbook is touchable.
- A rolling pre-write backup (master-tracker.pre-resolve-backup.xlsx) is saved
  next to the tracker before every write ("backup" in the name keeps it out of
  the reconciler's tracker picker).

Run:  python "Sch. E/_System/dashboard_server.py"  [--tracker <path>] [--port N]
Stop: Ctrl+C
"""
import argparse
import datetime
import json
import os
import shutil
import subprocess
import sys
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)                      # Sch. E folder
RECONCILE = os.path.join(HERE, "reconcile.py")
DASHBOARD = os.path.join(ROOT, "SchE_Dashboard.html")

sys.path.insert(0, HERE)
import reconcile as _rec                           # reuse its tracker discovery

TRACKER = _rec.TRACKER
PORT = 8742


def rerun_reconcile():
    """Regenerate dashboard/report from the books. Exit code = breaks (informational)."""
    subprocess.run([sys.executable, RECONCILE, ROOT, "--quiet"], cwd=os.path.dirname(ROOT))


def resolve_row(row, note):
    """Write Review!E<row> = Resolved. Refuses anything but an open Review row."""
    if not isinstance(row, int) or row < 5:
        return "invalid row"
    wb = openpyxl.load_workbook(TRACKER, data_only=False)
    if "Review" not in wb.sheetnames:
        return "no Review tab"
    ws = wb["Review"]
    if not ws.cell(row, 2).value:
        return f"Review row {row} has no item"
    current = str(ws.cell(row, 5).value or "")
    if current.lower().startswith("resolved"):
        return f"row {row} is already resolved"
    stamp = datetime.date.today().isoformat()
    text = f"Resolved — {note.strip()} ({stamp})" if note.strip() else f"Resolved ({stamp})"
    backup = os.path.join(os.path.dirname(TRACKER), "master-tracker.pre-resolve-backup.xlsx")
    shutil.copy2(TRACKER, backup)
    ws.cell(row, 5).value = text
    wb.save(TRACKER)
    return None  # success


class Handler(BaseHTTPRequestHandler):
    def _send(self, code, body, ctype="application/json"):
        data = body if isinstance(body, bytes) else json.dumps(body).encode()
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):
        if self.path in ("/", "/index.html", "/SchE_Dashboard.html"):
            with open(DASHBOARD, "rb") as f:
                self._send(200, f.read(), "text/html; charset=utf-8")
        else:
            self._send(404, {"error": "not found"})

    def do_POST(self):
        if self.path != "/api/resolve":
            return self._send(404, {"ok": False, "error": "unknown endpoint"})
        try:
            length = int(self.headers.get("Content-Length", 0))
            payload = json.loads(self.rfile.read(length) or b"{}")
            err = resolve_row(payload.get("row"), str(payload.get("note", "")))
        except Exception as e:  # malformed JSON, workbook locked (file open in Excel), etc.
            err = str(e)
        if err:
            return self._send(400, {"ok": False, "error": err})
        rerun_reconcile()
        self._send(200, {"ok": True})

    def log_message(self, fmt, *args):
        print("[dashboard]", fmt % args)


def main():
    global TRACKER, PORT
    ap = argparse.ArgumentParser()
    ap.add_argument("--tracker", help="override tracker path (testing)")
    ap.add_argument("--port", type=int, default=PORT)
    a = ap.parse_args()
    if a.tracker:
        TRACKER = os.path.abspath(a.tracker)
    PORT = a.port
    print(f"[dashboard] tracker: {TRACKER}")
    rerun_reconcile()  # fresh dashboard on startup
    url = f"http://127.0.0.1:{PORT}/"
    print(f"[dashboard] serving {url}  (Ctrl+C to stop)")
    ThreadingHTTPServer(("127.0.0.1", PORT), Handler).serve_forever()


if __name__ == "__main__":
    main()
