#!/usr/bin/env python3
"""Propose receipt-to-expense matches without changing the tracker or source files."""

from __future__ import annotations

import json
import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl


SKIP_NAMES = {
    "sche_dashboard.html",
    "sche_reconciliation.md",
    "reconciliation.json",
    "matchproposals.md",
    "match_state.json",
    "dashboard_state.json",
}
TOKEN_STOPWORDS = {
    "and",
    "for",
    "from",
    "invoice",
    "paid",
    "payment",
    "receipt",
    "replace",
    "repair",
    "sample",
    "the",
    "with",
}


def tracker_path(root: Path) -> Path | None:
    candidates = [
        path
        for path in root.glob("*.xlsx")
        if "backup" not in path.name.lower() and "test" not in path.name.lower()
    ]
    preferred = [path for path in candidates if "tracker" in path.name.lower()]
    return (preferred or candidates or [None])[0]


def read_profile(root: Path) -> dict:
    path = root / "profile.json"
    if not path.exists():
        return {}
    try:
        loaded = json.loads(path.read_text(encoding="utf-8"))
        return loaded if isinstance(loaded, dict) else {}
    except (OSError, ValueError):
        return {}


def candidate_files(root: Path, profile: dict, tracker: Path) -> list[Path]:
    roots = [root]
    for value in profile.get("document_folders", []):
        if isinstance(value, str) and value.strip():
            roots.append(Path(value).expanduser())

    found: dict[str, Path] = {}
    for base in roots:
        if not base.is_dir():
            continue
        for path in base.rglob("*"):
            if not path.is_file():
                continue
            try:
                resolved = path.resolve()
            except OSError:
                continue
            if resolved == tracker.resolve():
                continue
            if path.name.lower() in SKIP_NAMES:
                continue
            found[os.path.normcase(str(resolved))] = resolved
    return sorted(found.values(), key=lambda path: os.path.normcase(str(path)))


def row_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        match = re.search(r"(\d{4})[-_/](\d{2})[-_/](\d{2})", value)
        if match:
            try:
                return date(*(int(part) for part in match.groups()))
            except ValueError:
                pass
    return None


def file_date(path: Path) -> tuple[date | None, str]:
    name = path.name
    patterns = [
        r"(?<!\d)(\d{4})[-_](\d{2})[-_](\d{2})(?!\d)",
        r"(?<!\d)(\d{4})(\d{2})(\d{2})(?!\d)",
    ]
    for pattern in patterns:
        match = re.search(pattern, name)
        if match:
            try:
                return date(*(int(part) for part in match.groups())), "filename"
            except ValueError:
                pass
    try:
        return datetime.fromtimestamp(path.stat().st_mtime).date(), "modified time"
    except OSError:
        return None, "unavailable"


def amount_match(path: Path, amount: float) -> str | None:
    absolute = abs(amount)
    forms = {f"{absolute:.2f}"}
    if absolute.is_integer():
        forms.add(str(int(absolute)))
    for form in sorted(forms, key=len, reverse=True):
        pattern = rf"(?<!\d){re.escape(form)}(?!\d)"
        if re.search(pattern, path.stem.replace(",", "")):
            return form
    return None


def words(value: str) -> set[str]:
    return {
        token
        for token in re.findall(r"[a-z0-9]+", value.lower())
        if len(token) >= 3 and token not in TOKEN_STOPWORDS
    }


def describe_path(path: Path, root: Path) -> str:
    try:
        return path.relative_to(root).as_posix()
    except ValueError:
        return str(path)


def score_candidate(path: Path, expense: dict) -> dict | None:
    evidence = []
    score = 0

    candidate_date, date_source = file_date(path)
    if expense["date"] and candidate_date:
        distance = abs((candidate_date - expense["date"]).days)
        if distance <= 5:
            score += 1
            evidence.append(
                f"date {candidate_date.isoformat()} from {date_source} "
                f"is {distance} day(s) from row date"
            )

    matched_amount = amount_match(path, expense["amount"])
    if matched_amount:
        score += 1
        evidence.append(f"filename contains amount {matched_amount}")

    row_tokens = words(f"{expense['vendor']} {expense['description']}")
    filename_tokens = words(path.stem)
    token_hits = sorted(row_tokens & filename_tokens)
    if token_hits:
        score += 1
        evidence.append("filename contains row token(s): " + ", ".join(token_hits))

    if score < 2:
        return None
    return {
        "path": path,
        "score": score,
        "evidence": evidence,
    }


def read_unlinked_expenses(tracker: Path) -> list[dict]:
    workbook = openpyxl.load_workbook(
        tracker, read_only=True, data_only=False
    )
    try:
        sheet = workbook["Expenses"]
        rows = []
        for row in range(5, sheet.max_row + 1):
            vendor = sheet.cell(row, 2).value
            date_value = sheet.cell(row, 1).value
            if (vendor in (None, "TOTAL")) and date_value in (None, "TOTAL"):
                continue
            if str(sheet.cell(row, 9).value or "").strip():
                continue
            paid = sheet.cell(row, 6).value
            contracted = sheet.cell(row, 5).value
            try:
                amount = float(paid if paid not in (None, "") else contracted or 0)
            except (TypeError, ValueError):
                amount = 0.0
            rows.append(
                {
                    "row": row,
                    "date": row_date(date_value),
                    "vendor": str(vendor or ""),
                    "description": str(sheet.cell(row, 3).value or ""),
                    "amount": amount,
                }
            )
        return rows
    finally:
        workbook.close()


def previous_misses(root: Path) -> dict[str, dict]:
    path = root / "match_state.json"
    if not path.exists():
        return {}
    try:
        state = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return {}
    return {
        str(item.get("row")): item
        for item in state.get("misses", [])
        if isinstance(item, dict) and item.get("row") is not None
    }


def expense_record(expense: dict) -> dict:
    return {
        "row": expense["row"],
        "vendor": expense["vendor"],
        "date": expense["date"].isoformat() if expense["date"] else "",
        "amount": round(expense["amount"], 2),
    }


def build_state(root: Path, tracker: Path) -> dict:
    profile = read_profile(root)
    candidates = candidate_files(root, profile, tracker)
    old_misses = previous_misses(root)
    proposals = []
    ambiguous = []
    misses = []

    for expense in read_unlinked_expenses(tracker):
        matches = [
            match
            for path in candidates
            if (match := score_candidate(path, expense)) is not None
        ]
        base = expense_record(expense)
        if len(matches) == 1:
            match = matches[0]
            proposals.append(
                {
                    **base,
                    "file": describe_path(match["path"], root),
                    "paste_value": describe_path(match["path"], root),
                    "score": match["score"],
                    "evidence": match["evidence"],
                }
            )
        elif len(matches) > 1:
            ambiguous.append(
                {
                    **base,
                    "candidates": [
                        {
                            "file": describe_path(match["path"], root),
                            "score": match["score"],
                            "evidence": match["evidence"],
                        }
                        for match in matches
                    ],
                }
            )
        else:
            prior = old_misses.get(str(expense["row"]), {})
            same_item = (
                prior.get("vendor") == expense["vendor"]
                and prior.get("date") == base["date"]
                and prior.get("amount") == base["amount"]
            )
            misses.append(
                {
                    **base,
                    "miss_count": int(prior.get("miss_count", 0)) + 1 if same_item else 1,
                }
            )

    return {
        "version": 1,
        "tracker": tracker.name,
        "proposals": proposals,
        "ambiguous": ambiguous,
        "misses": misses,
    }


def money(value: float) -> str:
    return f"${value:,.2f}"


def write_proposals(root: Path, state: dict) -> None:
    lines = [
        "# Receipt match proposals",
        "",
        "_Proposals only. This tool did not change the tracker or any source file._",
        "",
    ]
    need = [item for item in state["misses"] if item["miss_count"] >= 2]
    if need:
        lines.extend(["## Need from you", ""])
        for item in need:
            lines.append(
                f"- Expenses row {item['row']}: **{item['vendor']}** · "
                f"{item['date'] or 'date unknown'} · {money(item['amount'])} "
                f"({item['miss_count']} searches with no match)"
            )
        lines.append("")

    lines.extend(["## Proposals", ""])
    if not state["proposals"]:
        lines.append("_None._")
    for item in state["proposals"]:
        lines.extend(
            [
                f"### Expenses row {item['row']} — {item['vendor']} — {money(item['amount'])}",
                "",
                f"- Proposed file: `{item['file']}`",
                "- Evidence: " + "; ".join(item["evidence"]),
                f"- Human action: paste `{item['paste_value']}` into the File Reference cell.",
                "",
            ]
        )

    lines.extend(["## Ambiguous", ""])
    if not state["ambiguous"]:
        lines.append("_None._")
    for item in state["ambiguous"]:
        lines.append(
            f"### Expenses row {item['row']} — {item['vendor']} — {money(item['amount'])}"
        )
        lines.append("")
        lines.append("No proposal was made because more than one candidate is plausible:")
        for candidate in item["candidates"]:
            lines.append(
                f"- `{candidate['file']}` — " + "; ".join(candidate["evidence"])
            )
        lines.append("")

    waiting = [item for item in state["misses"] if item["miss_count"] < 2]
    lines.extend(["## Not found yet", ""])
    if not waiting:
        lines.append("_None._")
    for item in waiting:
        lines.append(
            f"- Expenses row {item['row']}: **{item['vendor']}** · "
            f"{item['date'] or 'date unknown'} · {money(item['amount'])} "
            f"(miss {item['miss_count']} of 2)"
        )
    lines.append("")
    (root / "MatchProposals.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    roots = [arg for arg in sys.argv[1:] if not arg.startswith("--")]
    root = Path(roots[0] if roots else "fixtures/sample-property").resolve()
    tracker = tracker_path(root)
    if tracker is None or not tracker.exists():
        print(f"Tracker not found in: {root}")
        return 2

    state = build_state(root, tracker)
    (root / "match_state.json").write_text(
        json.dumps(state, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    write_proposals(root, state)
    print(
        f"Receipt matcher: {len(state['proposals'])} proposal(s), "
        f"{len(state['ambiguous'])} ambiguous, {len(state['misses'])} miss(es)"
    )
    print(f"  Proposals: {root / 'MatchProposals.md'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
