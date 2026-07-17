#!/usr/bin/env python3
"""Generate the sanitized 12 Sample Street demonstration fixture."""

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


HERE = Path(__file__).resolve().parent
PROPERTY = HERE / "sample-property"
WORKBOOK = PROPERTY / "sample-tracker.xlsx"
SOURCE_DOCS = PROPERTY / "source-docs"

INK = "2B2A26"
MUTED = "7D7666"
CREAM = "F4EDDD"
CARD = "FBF8EF"
OLIVE = "566123"
CRIMSON = "9B1C2E"
SAND = "D2C8B0"
WHITE = "FFFFFF"
THIN = Side(style="thin", color=SAND)


def write_placeholder_pdf(path: Path, title: str) -> None:
    """Write a small, valid one-page PDF using only the Python standard library."""
    safe_title = title.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
    stream = f"BT /F1 14 Tf 72 720 Td ({safe_title}) Tj ET".encode("ascii")
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
        b"/Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n"
        + stream + b"\nendstream",
    ]
    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for number, obj in enumerate(objects, start=1):
        offsets.append(len(pdf))
        pdf.extend(f"{number} 0 obj\n".encode("ascii"))
        pdf.extend(obj)
        pdf.extend(b"\nendobj\n")
    xref = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    pdf.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    pdf.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
        f"startxref\n{xref}\n%%EOF\n".encode("ascii")
    )
    path.write_bytes(pdf)


def set_widths(ws, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def style_title(ws, title: str, subtitle: str, end_column: str) -> None:
    ws.merge_cells(f"A1:{end_column}1")
    ws["A1"] = title
    ws["A1"].fill = PatternFill("solid", fgColor=OLIVE)
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=WHITE)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A2:{end_column}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=MUTED)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 28


def style_header(ws, row: int, start: int, end: int) -> None:
    for cell in ws.iter_cols(min_col=start, max_col=end, min_row=row, max_row=row):
        c = cell[0]
        c.fill = PatternFill("solid", fgColor=INK)
        c.font = Font(name="Calibri", bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=THIN)
    ws.row_dimensions[row].height = 28


def style_body(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in ws.iter_rows(
        min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col
    ):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=CARD)
            cell.font = Font(name="Calibri", size=10, color=INK)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)


def build_income(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Income"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    style_title(
        ws,
        "12 Sample Street — 2026 Income",
        "Fake demonstration data. January–April are actual; May–December are projections.",
        "E",
    )
    ws.append([])
    ws.append(["Property", "12 Sample Street", "Tax year", 2026, "Cash basis"])
    ws.append(["Month", "Gross Rent", "Tenant Credit", "Net Rent", "Status"])
    style_header(ws, 5, 1, 5)

    rents = [
        ("January", 1800, 0, "Actual"),
        ("February", 1800, 50, "Actual"),
        ("March", 1800, 0, "Actual"),
        ("April", 1800, 0, "Actual"),
        ("May", 1800, 0, "Projected"),
        ("June", 1800, 0, "Projected"),
        ("July", 1800, 0, "Projected"),
        ("August", 1800, 0, "Projected"),
        ("September", 1800, 0, "Projected"),
        ("October", 1800, 0, "Projected"),
        ("November", 1800, 0, "Projected"),
        ("December", 1800, 0, "Projected"),
    ]
    for month, rent, credit, status in rents:
        ws.append([month, rent, credit, rent - credit, status])
    style_body(ws, 6, 17, 1, 5)
    for row in range(6, 18):
        for column in (2, 3, 4):
            ws.cell(row, column).number_format = '$#,##0.00;[Red]($#,##0.00);-'
            ws.cell(row, column).alignment = Alignment(horizontal="right")

    ws["A19"] = "YTD net income (Jan–Apr actual)"
    ws["D19"] = sum(rent - credit for _, rent, credit, status in rents if status == "Actual")
    for cell in ws[19]:
        cell.fill = PatternFill("solid", fgColor=CREAM)
        cell.font = Font(name="Calibri", bold=True, color=INK)
        cell.border = Border(top=Side(style="medium", color=OLIVE))
    ws["D19"].number_format = '$#,##0.00;[Red]($#,##0.00);-'

    status_validation = DataValidation(
        type="list", formula1='"Actual,Projected"', allow_blank=False
    )
    ws.add_data_validation(status_validation)
    status_validation.add("E6:E17")
    set_widths(ws, {"A": 18, "B": 15, "C": 15, "D": 15, "E": 14})


def build_expenses(wb: Workbook) -> None:
    ws = wb.create_sheet("Expenses")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    style_title(
        ws,
        "12 Sample Street — 2026 Expenses",
        "Fake source references are relative to this fixture folder. Two empty references test receipt matching.",
        "J",
    )
    ws.append(["Property", "12 Sample Street", "Tax year", 2026])
    ws.append(
        [
            "Date",
            "Vendor",
            "Description",
            "Schedule E line",
            "Contracted",
            "Paid YTD",
            "Balance",
            "Status",
            "File Reference",
            "Notes",
        ]
    )
    style_header(ws, 4, 1, 10)

    rows = [
        [
            date(2026, 1, 10),
            "Sample Mutual",
            "Annual landlord policy",
            "L9 — Insurance",
            1200,
            1200,
            0,
            "Paid",
            "source-docs/sample-insurance-policy.txt",
            "Fixture policy record.",
        ],
        [
            date(2026, 1, 22),
            "Example Tax Services",
            "Bookkeeping setup consultation",
            "L10 — Professional fees",
            350,
            350,
            0,
            "Paid",
            "source-docs/bookkeeping-consult.txt",
            "Fixture professional-fee record.",
        ],
        [
            date(2026, 2, 8),
            "Sample Plumber",
            "Repair leaking kitchen supply line — Claim SAMPLE-001",
            "L14 — Repairs",
            625,
            625,
            0,
            "Paid",
            "source-docs/plumbing-invoice.pdf",
            "Claim SAMPLE-001 repair invoice.",
        ],
        [
            date(2026, 2, 12),
            "Anytown Hardware",
            "Paint and patch supplies",
            "L14 — Repairs",
            210,
            210,
            0,
            "Paid",
            "source-docs/hardware-receipt.txt",
            "Fixture repair supplies.",
        ],
        [
            date(2026, 3, 1),
            "Anytown County",
            "First property-tax installment",
            "L16 — Taxes",
            950,
            950,
            0,
            "Paid",
            "source-docs/property-tax-receipt.pdf",
            "Fixture county receipt.",
        ],
        [
            date(2026, 3, 15),
            "Anytown County",
            "Rental registration assessment",
            "L16 — Taxes",
            400,
            400,
            0,
            "Paid",
            "source-docs/registration-assessment.txt",
            "Fixture local assessment.",
        ],
        [
            date(2026, 3, 20),
            "Sample Mutual",
            "Insurance reimbursement proceeds — Claim SAMPLE-001",
            "L14 — Repairs",
            -300,
            -300,
            0,
            "Received",
            "source-docs/reimbursement-notice.pdf",
            "ACCOUNTANT REVIEW: Confirm the negative repair contra treatment with the preparer.",
        ],
        [
            date(2026, 4, 20),
            "Sample Gardener",
            "Spring yard cleanup",
            "L14 — Repairs",
            180,
            180,
            0,
            "Paid",
            "source-docs/missing-gardener-invoice.pdf",
            "Intentionally dangling reference — demonstrates the reconciler catching a broken link.",
        ],
        [
            date(2026, 4, 5),
            "Sample Locksmith",
            "Replace rear-entry lock",
            "L14 — Repairs",
            95,
            95,
            0,
            "Paid",
            "",
            "Intentionally unlinked; its receipt genuinely does not exist.",
        ],
        [
            date(2026, 4, 10),
            "Sample Electric",
            "Replace porch light fixture",
            "L14 — Repairs",
            145,
            145,
            0,
            "Paid",
            "",
            "Intentionally unlinked; its matching receipt exists in source-docs.",
        ],
    ]
    for row in rows:
        ws.append(row)
    style_body(ws, 5, 13, 1, 10)
    for row in range(5, 14):
        ws.cell(row, 1).number_format = "yyyy-mm-dd"
        for column in (5, 6, 7):
            ws.cell(row, column).number_format = '$#,##0.00;[Red]($#,##0.00);-'
            ws.cell(row, column).alignment = Alignment(horizontal="right")
    ws.conditional_formatting.add(
        "J5:J13",
        FormulaRule(
            formula=['ISNUMBER(SEARCH("ACCOUNTANT REVIEW",J5))'],
            fill=PatternFill("solid", fgColor="FFF2CC"),
            font=Font(color=CRIMSON, bold=True),
        ),
    )
    set_widths(
        ws,
        {
            "A": 13,
            "B": 20,
            "C": 34,
            "D": 28,
            "E": 14,
            "F": 14,
            "G": 14,
            "H": 13,
            "I": 42,
            "J": 58,
        },
    )


def build_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("Sch. E Summary")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B7"
    style_title(
        ws,
        "12 Sample Street — Schedule E Summary",
        "Formulas only on this tab. The reconciler recomputes all totals from Income and Expenses.",
        "D",
    )
    ws["A4"] = "Property"
    ws["B4"] = "12 Sample Street"
    ws["A5"] = "Fair Rental Days"
    ws["B5"] = 365
    ws.append([])
    ws.append(["", "Schedule E line", "Amount", "Notes"])
    style_header(ws, 7, 2, 4)

    categories = [
        (5, "Advertising"),
        (6, "Auto and travel"),
        (7, "Cleaning and maintenance"),
        (8, "Commissions"),
        (9, "Insurance"),
        (10, "Professional fees"),
        (11, "Management fees"),
        (12, "Mortgage interest"),
        (13, "Other interest"),
        (14, "Repairs"),
        (15, "Supplies"),
        (16, "Taxes"),
        (17, "Utilities"),
        (18, "Depreciation"),
        (19, "Other"),
    ]
    row = 8
    ws.cell(row, 2, "3 — Rents received")
    ws.cell(row, 3, "='Income'!D19")
    ws.cell(row, 4, "Four actual months.")
    row += 1
    ws.cell(row, 2, "4 — Royalties")
    ws.cell(row, 3, "=0")
    ws.cell(row, 4, "Not applicable to this fixture.")
    for line, label in categories:
        row += 1
        exact = f"L{line} — {label}"
        ws.cell(row, 2, f"{line} — {label}")
        ws.cell(
            row,
            3,
            f'=SUMIF(\'Expenses\'!D:D,"{exact}",\'Expenses\'!F:F)',
        )
        ws.cell(row, 4, f"Rolls up exact label: {exact}")
    row += 1
    total_row = row
    ws.cell(row, 2, "20 — Total expenses")
    ws.cell(row, 3, f"=SUM(C10:C{row - 1})")
    ws.cell(row, 4, "Sum of lines 5–19.")
    row += 1
    ws.cell(row, 2, "21 — Net income")
    ws.cell(row, 3, f"=C8+C9-C{total_row}")
    ws.cell(row, 4, "Line 3 + Line 4 − Line 20.")

    style_body(ws, 8, row, 2, 4)
    for r in range(8, row + 1):
        ws.cell(r, 3).number_format = '$#,##0.00;[Red]($#,##0.00);-'
        ws.cell(r, 3).alignment = Alignment(horizontal="right")
    for r in (total_row, row):
        for cell in ws[r]:
            cell.font = Font(name="Calibri", bold=True, color=INK)
            cell.border = Border(top=Side(style="medium", color=OLIVE))
    set_widths(ws, {"A": 4, "B": 31, "C": 18, "D": 54})


def build_review(wb: Workbook) -> None:
    ws = wb.create_sheet("Review")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    style_title(
        ws,
        "12 Sample Street — Review Queue",
        "Three open and two resolved fake demonstration items.",
        "E",
    )
    ws.append(["Queue", "Sample review items"])
    ws.append(["Type", "Item", "Issue", "Action Needed", "Status"])
    style_header(ws, 4, 1, 5)
    rows = [
        [
            "Missing data",
            "Sample Locksmith receipt",
            "No matching receipt is available for the expense row.",
            "Locate the receipt and paste its path into File Reference.",
            "Open",
        ],
        [
            "Insurance allocation",
            "Reimbursement received $300.00",
            "Claim SAMPLE-001 proceeds are booked as a negative Line 14 contra row.",
            "Ask the preparer to confirm the treatment.",
            "Open",
        ],
        [
            "Filing question",
            "Fair rental days",
            "The sample workbook uses 365 days.",
            "Confirm the day count before filing.",
            "Open",
        ],
        [
            "Document",
            "2026 sample lease",
            "Fixture lease copy was checked.",
            "No action needed.",
            "Resolved — fixture review complete",
        ],
        [
            "Income",
            "January rent deposit",
            "Fixture deposit matches the income row.",
            "No action needed.",
            "Resolved — fixture review complete",
        ],
    ]
    for review_row in rows:
        ws.append(review_row)
    style_body(ws, 5, 9, 1, 5)
    status_validation = DataValidation(
        type="list",
        formula1='"Open,Resolved — fixture review complete"',
        allow_blank=False,
    )
    ws.add_data_validation(status_validation)
    status_validation.add("E5:E9")
    ws.conditional_formatting.add(
        "E5:E9",
        FormulaRule(
            formula=['LEFT(E5,8)="Resolved"'],
            fill=PatternFill("solid", fgColor="E2F0D9"),
            font=Font(color=OLIVE),
        ),
    )
    set_widths(ws, {"A": 22, "B": 30, "C": 48, "D": 44, "E": 34})


def write_source_documents() -> None:
    SOURCE_DOCS.mkdir(parents=True, exist_ok=True)
    text_documents = {
        "sample-insurance-policy.txt": (
            "FAKE DEMONSTRATION DOCUMENT\n"
            "Property: 12 Sample Street, Anytown, CA 00000\n"
            "Annual landlord policy premium: $1,200.00\n"
        ),
        "bookkeeping-consult.txt": (
            "FAKE DEMONSTRATION DOCUMENT\n"
            "Example Tax Services bookkeeping setup: $350.00\n"
        ),
        "hardware-receipt.txt": (
            "FAKE DEMONSTRATION DOCUMENT\n"
            "Paint and patch supplies: $210.00\n"
        ),
        "registration-assessment.txt": (
            "FAKE DEMONSTRATION DOCUMENT\n"
            "Anytown rental registration assessment: $400.00\n"
        ),
    }
    for name, content in text_documents.items():
        (SOURCE_DOCS / name).write_text(content, encoding="utf-8")
    write_placeholder_pdf(
        SOURCE_DOCS / "plumbing-invoice.pdf",
        "FAKE: Sample Plumber invoice - 12 Sample Street - $625.00",
    )
    write_placeholder_pdf(
        SOURCE_DOCS / "property-tax-receipt.pdf",
        "FAKE: Anytown County property tax receipt - $950.00",
    )
    write_placeholder_pdf(
        SOURCE_DOCS / "reimbursement-notice.pdf",
        "FAKE: Claim SAMPLE-001 reimbursement received - $300.00",
    )
    write_placeholder_pdf(
        SOURCE_DOCS / "2026-04-10_sample-electric_145.pdf",
        "FAKE: Sample Electric receipt - 12 Sample Street - $145.00",
    )
    dangling = SOURCE_DOCS / "missing-locksmith-receipt.pdf"
    if dangling.exists():
        dangling.unlink()


def build_workbook() -> None:
    PROPERTY.mkdir(parents=True, exist_ok=True)
    write_source_documents()
    wb = Workbook()
    build_income(wb)
    build_expenses(wb)
    build_summary(wb)
    build_review(wb)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(WORKBOOK)


def build_empty_workbook(
    property_name: str,
    output_path: Path,
    applicable_lines: list[int],
    tracker_type: str = "rental",
) -> None:
    """Create an empty tracker using the demonstration workbook's four-tab layout."""
    wb = Workbook()
    build_income(wb)
    build_expenses(wb)
    build_summary(wb)
    build_review(wb)

    schedule = "Schedule E" if tracker_type == "rental" else "Schedule C setup preview"

    income = wb["Income"]
    income["A1"] = f"{property_name} — Income"
    income["A2"] = "Empty tracker. Add income only after confirming it from source records."
    income["B4"] = property_name
    for row in range(6, 18):
        for column in range(2, 6):
            income.cell(row, column).value = None
    income["A19"] = "YTD net income from Actual months"
    income["D19"] = 0

    expenses = wb["Expenses"]
    expenses["A1"] = f"{property_name} — Expenses"
    expenses["A2"] = "Empty tracker. File references are relative to this tracker folder."
    expenses["B3"] = property_name
    expenses.delete_rows(5, max(expenses.max_row - 4, 0))
    style_body(expenses, 5, 24, 1, 10)
    expense_status = DataValidation(
        type="list", formula1='"Paid,Pending,Received,TBD"', allow_blank=True
    )
    expenses.add_data_validation(expense_status)
    expense_status.add("H5:H104")

    summary = wb["Sch. E Summary"]
    summary["A1"] = f"{property_name} — {schedule} Summary"
    summary["A2"] = (
        "Formulas only on this tab. The reconciler recomputes totals from Income and Expenses."
    )
    summary["B4"] = property_name
    summary["B5"] = None
    summary["D8"] = "Actual months only."
    summary["D9"] = "Leave at zero unless applicable."
    summary.column_dimensions["A"].width = 16
    applicable = set(applicable_lines)
    for row in range(10, 25):
        line = row - 5
        if line not in applicable:
            for cell in summary[row][1:4]:
                cell.fill = PatternFill("solid", fgColor="E6E2D8")
                cell.font = Font(name="Calibri", size=10, color=MUTED)
            summary.cell(row, 4).value = "Not selected during setup."

    review = wb["Review"]
    review["A1"] = f"{property_name} — Review Queue"
    review["A2"] = "Empty queue. Add questions or missing-data items for review."
    review["B3"] = "Setup complete"
    review.delete_rows(5, max(review.max_row - 4, 0))
    style_body(review, 5, 24, 1, 5)
    review_status = DataValidation(
        type="list", formula1='"Open,Resolved"', allow_blank=True
    )
    review.add_data_validation(review_status)
    review_status.add("E5:E104")

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    build_workbook()
    print(f"Created fake sample fixture: {WORKBOOK}")
